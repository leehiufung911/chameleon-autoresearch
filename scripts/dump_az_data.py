"""Dump full AZ 2026 data from both sheets."""
import openpyxl
import csv

wb = openpyxl.load_workbook(
    'C:/Users/mic23/prototype-embed/chameleon-research-loop/papers/AZ2026/ml6c00043_si_002.xlsx'
)

# Sheet 1: Reference compounds
print("=== REFERENCE COMPOUNDS ===")
ws = wb['Reference compounds']
for row in ws.iter_rows(values_only=True):
    if row[0] is not None:
        print(f"  {row[0]:<20} ChamelogK={row[2]}  (lit={row[4]})")

# Sheet 2: 68 PROTACs - dump to TSV
print("\n=== 68 PROTACs ===")
ws = wb['68 PROTACs']
headers = None
rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        headers = list(row)
        continue
    if row[0] is None:
        continue
    rows.append(list(row))

# Print summary
print(f"  {len(rows)} compounds with data")
print(f"  Columns: {headers}")

# Print all compound IDs with key data
print("\n  ID | MW | Chamelogk100 | ePSA | TPSA | ETR | Mouse_BA% | mFabs")
print("  " + "-"*80)
for r in rows:
    cid = r[0] if r[0] else "?"
    mw = f"{r[1]:.0f}" if r[1] else "?"
    clk = f"{r[4]:.2f}" if r[4] else "?"
    epsa = f"{r[9]:.0f}" if r[9] else "?"
    tpsa = f"{r[10]:.0f}" if r[10] else "?"
    etr = f"{r[11]:.2f}" if r[11] else "?"
    ba = f"{r[14]:.1f}" if r[14] else "?"
    mfabs = f"{r[15]:.2f}" if r[15] else "?"
    print(f"  {cid:<20} {mw:>5} {clk:>8} {epsa:>6} {tpsa:>6} {etr:>6} {ba:>8} {mfabs:>6}")

# Save to TSV
outpath = 'C:/Users/mic23/prototype-embed/chameleon-research-loop/papers/AZ2026/az_68_protacs.tsv'
with open(outpath, 'w', newline='') as f:
    w = csv.writer(f, delimiter='\t')
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
print(f"\nSaved to {outpath}")
