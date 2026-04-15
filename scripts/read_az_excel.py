"""Read AZ 2026 SI Excel file and dump contents."""
import openpyxl
import sys

wb = openpyxl.load_workbook(
    'C:/Users/mic23/prototype-embed/chameleon-research-loop/papers/AZ2026/ml6c00043_si_002.xlsx'
)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f'=== Sheet: {sheet} ({ws.max_row} rows x {ws.max_column} cols) ===')
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5 or i == ws.max_row - 1:
            print(f"  Row {i}: {list(row)}")
        elif i == 5:
            print(f"  ... ({ws.max_row - 6} more rows) ...")
    print()
